"""
Online Excell v2.1 — Premium Follow-up Tracker
Cache-first architecture: in-memory JSON sheets, async Turso persistence.
Multi-file upload with drag & drop, supports .xlsx/.xlsm/.xls/.csv.
"""

import os, re, sqlite3, json, io, csv, logging, threading
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from dateutil import parser as dateparser
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, StreamingResponse
import requests as http_requests
from apscheduler.schedulers.background import BackgroundScheduler

log = logging.getLogger("excell")
app = FastAPI(title="Online Excell", version="2.1.0")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Cloud Config ──────────────────────────────────────────────────────────────
TURSO_URL = os.environ.get("TURSO_URL", "")
TURSO_AUTH_TOKEN = os.environ.get("TURSO_AUTH_TOKEN", "")
USE_TURSO = bool(TURSO_URL and TURSO_AUTH_TOKEN)
COL_EXTRA = 20

# ══════════════════════════════════════════════════════════════════════════════
#  IN-MEMORY CACHE  — the heart of the speed system
# ══════════════════════════════════════════════════════════════════════════════
CACHE = {
    "sheets": {},        # "2026_6" -> {"month":6,"year":2026,"entries":[...],...}
    "col_names": {},     # "col1" -> "Note 1", ...
    "master_count": 0,
    "_id_counter": 0,    # global auto-increment for entry IDs
    "_loaded": False,
}
_cache_lock = threading.Lock()

def sheet_key(year, month):
    return f"{year}_{month}"

def next_id():
    with _cache_lock:
        CACHE["_id_counter"] += 1
        return CACHE["_id_counter"]


# ══════════════════════════════════════════════════════════════════════════════
#  TURSO HTTP CLIENT  — minimal, only for persistence
# ══════════════════════════════════════════════════════════════════════════════

class TursoClient:
    """Thin HTTP client for Turso — sends SQL over the v2 pipeline API."""
    def __init__(self, url, token):
        self.endpoint = url.replace("libsql://", "https://").rstrip("/") + "/v2/pipeline"
        self.headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    @staticmethod
    def _typed(val):
        if val is None: return {"type": "null"}
        if isinstance(val, bool): return {"type": "integer", "value": str(int(val))}
        if isinstance(val, int): return {"type": "integer", "value": str(val)}
        if isinstance(val, float): return {"type": "float", "value": str(val)}
        return {"type": "text", "value": str(val)}

    @staticmethod
    def _extract(v):
        if v is None or v.get("type") == "null": return None
        if v.get("type") == "integer": return int(v["value"])
        if v.get("type") == "float": return float(v["value"])
        return v.get("value", "")

    def execute(self, sql, params=None):
        """Single query → returns list of dicts."""
        stmt = {"sql": sql}
        if params:
            stmt["args"] = [self._typed(p) for p in params]
        body = {"requests": [{"type": "execute", "stmt": stmt}, {"type": "close"}]}
        resp = http_requests.post(self.endpoint, json=body, headers=self.headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        res = data["results"][0]
        if res["type"] == "error":
            raise Exception(f"Turso: {res['error'].get('message', res['error'])}")
        result = res["response"]["result"]
        cols = [c["name"] for c in result.get("cols", [])]
        rows = []
        for raw_row in result.get("rows", []):
            row = {}
            for i, val in enumerate(raw_row):
                col_name = cols[i] if i < len(cols) else f"col{i}"
                row[col_name] = self._extract(val)
            rows.append(row)
        return rows

    def pipeline(self, queries):
        """Multiple queries in ONE HTTP call → returns list of list-of-dicts."""
        reqs = []
        for sql, params in queries:
            stmt = {"sql": sql}
            if params:
                stmt["args"] = [self._typed(p) for p in params]
            reqs.append({"type": "execute", "stmt": stmt})
        reqs.append({"type": "close"})
        resp = http_requests.post(self.endpoint, json={"requests": reqs},
                                  headers=self.headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        results = []
        for res in data["results"]:
            if res.get("type") == "error":
                raise Exception(f"Turso pipeline: {res['error']}")
            if res.get("type") == "ok":
                result = res["response"]["result"]
                cols = [c["name"] for c in result.get("cols", [])]
                rows = []
                for raw_row in result.get("rows", []):
                    row = {}
                    for i, val in enumerate(raw_row):
                        col_name = cols[i] if i < len(cols) else f"col{i}"
                        row[col_name] = self._extract(val)
                    rows.append(row)
                results.append(rows)
        return results

    def batch_upsert(self, statements):
        """Batch write — chunked into 200-statement groups."""
        chunk_size = 200
        reqs = []
        for sql, params in statements:
            stmt = {"sql": sql}
            if params:
                stmt["args"] = [self._typed(p) for p in params]
            reqs.append({"type": "execute", "stmt": stmt})

        for i in range(0, len(reqs), chunk_size):
            chunk = reqs[i:i + chunk_size] + [{"type": "close"}]
            resp = http_requests.post(self.endpoint, json={"requests": chunk},
                                      headers=self.headers, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            for res in data["results"]:
                if res.get("type") == "error":
                    raise Exception(f"Turso batch: {res['error']}")


# ── Local SQLite (dev fallback) ───────────────────────────────────────────────

DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "data.db"))

def _dict_factory(cursor, row):
    return dict(zip([col[0] for col in cursor.description], row))

def local_execute(sql, params=None):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _dict_factory
    conn.execute("PRAGMA journal_mode=WAL")
    cur = conn.execute(sql, params or ())
    rows = cur.fetchall()
    conn.commit()
    conn.close()
    return rows

def local_execute_many(statements):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _dict_factory
    conn.execute("PRAGMA journal_mode=WAL")
    results = []
    for sql, params in statements:
        cur = conn.execute(sql, params or ())
        results.append(cur.fetchall())
    conn.commit()
    conn.close()
    return results


# ── Unified DB interface ──────────────────────────────────────────────────────

turso = TursoClient(TURSO_URL, TURSO_AUTH_TOKEN) if USE_TURSO else None

def db_exec(sql, params=None):
    if USE_TURSO:
        return turso.execute(sql, params)
    return local_execute(sql, params)

def db_pipeline(queries):
    if USE_TURSO:
        return turso.pipeline(queries)
    return local_execute_many(queries)

def db_batch(statements):
    if USE_TURSO:
        return turso.batch_upsert(statements)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    for sql, params in statements:
        conn.execute(sql, params or ())
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════════════════════════
#  COLUMN MAPPING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

COLUMN_ALIASES = {
    'policyno': ['policy no', 'pol no', 'policy number', 'policyno', 'polno',
                 'pol number', 'policy_no', 'policy'],
    'name':     ['name', 'policyholder', 'holder name', 'insured name',
                 'policy holder', 'proposer name', 'proposer', 'life assured', 'party name'],
    'doc':      ['doc', 'date of commencement', 'commencement date', 'comm date',
                 'commencement', 'date of comm', 'commence date'],
    'plan':     ['plan', 'plan code', 'plan no', 'plancode', 'plan number',
                 'plan name', 'table term'],
    'mode':     ['mode', 'payment mode', 'pmt mode', 'pay mode', 'frequency',
                 'premium mode'],
    'premium':  ['premium', 'prem', 'premium amount', 'prem amt',
                 'installment premium', 'inst premium', 'modal premium'],
    'sumass':   ['sum assured', 'sum ass', 'sumass', 'sa', 'sum_assured',
                 'sumassured', 'sum assd', 'basic sa'],
    'mobileno': ['mobile', 'mobile no', 'mob', 'phone', 'contact',
                 'mobile number', 'mob no', 'phone no', 'contact no', 'cell', 'whatsapp'],
    'status':   ['status', 'payment status', 'sts', 'pay status', 'policy status'],
}

def _norm(s):
    return re.sub(r'[^a-z0-9 ]', '', str(s).lower().strip()).strip()

def map_columns(headers):
    mapping = {}
    used = set()
    normed_aliases = {}
    for field, aliases in COLUMN_ALIASES.items():
        normed_aliases[field] = [_norm(a) for a in aliases] + [_norm(field)]
    for i, raw in enumerate(headers):
        if raw is None: continue
        norm = _norm(raw)
        if not norm: continue
        for field, norms in normed_aliases.items():
            if field in used: continue
            if norm in norms or any(n in norm for n in norms if len(n) >= 3):
                mapping[i] = field
                used.add(field)
                break
    return mapping


# ── File Parsers ──────────────────────────────────────────────────────────────

def parse_csv_file(content: bytes) -> list:
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode('latin-1', errors='replace')
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(cell.strip() for cell in row)]

def parse_excel_file(content: bytes) -> list:
    """Parse .xlsx and .xlsm files using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        r = [str(v).strip() if v is not None else '' for v in row]
        if any(r):
            rows.append(r)
    wb.close()
    return rows

def parse_xls_file(content: bytes) -> list:
    """Parse legacy .xls files using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    rows = []
    for rx in range(ws.nrows):
        r = [str(ws.cell_value(rx, cx)).strip() for cx in range(ws.ncols)]
        if any(r):
            rows.append(r)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  MODE / FUP ENGINE
# ══════════════════════════════════════════════════════════════════════════════

MODE_MAP = {
    "m": 1, "mly": 1, "monthly": 1,
    "q": 3, "qly": 3, "quarterly": 3,
    "h": 6, "hly": 6, "halfyearly": 6, "half yearly": 6, "hy": 6,
    "a": 12, "ann": 12, "y": 12, "yly": 12, "yearly": 12, "annual": 12, "annually": 12,
}

def get_mode_months(mode_str):
    if not mode_str: return None
    cleaned = re.sub(r"[^a-z ]", "", str(mode_str).lower().strip())
    return MODE_MAP.get(cleaned) or MODE_MAP.get(cleaned.replace(" ", ""))

def parse_doc(doc_str):
    if not doc_str: return None
    try:
        dt = dateparser.parse(str(doc_str), dayfirst=True)
        return dt.date() if dt else None
    except:
        return None

def is_due_in_month(doc_str, mode_str, year, month):
    doc_date = parse_doc(doc_str)
    interval = get_mode_months(mode_str)
    if not doc_date or not interval: return False
    target_start = date(year, month, 1)
    target_end = date(year + (1 if month == 12 else 0), (month % 12) + 1, 1)
    if doc_date >= target_end: return False
    current = doc_date
    if current < target_start:
        months_diff = (target_start.year - current.year) * 12 + (target_start.month - current.month)
        skip = max(0, (months_diff // interval) - 1)
        if skip > 0:
            current = doc_date + relativedelta(months=interval * skip)
    while current < target_end:
        if target_start <= current < target_end: return True
        current += relativedelta(months=interval)
        if current.year > year + 2: break
    return False

def get_due_date(doc_str, mode_str, year, month):
    doc_date = parse_doc(doc_str)
    interval = get_mode_months(mode_str)
    if not doc_date or not interval: return None
    target_start = date(year, month, 1)
    target_end = date(year + (1 if month == 12 else 0), (month % 12) + 1, 1)
    current = doc_date
    if current < target_start:
        months_diff = (target_start.year - current.year) * 12 + (target_start.month - current.month)
        skip = max(0, (months_diff // interval) - 1)
        if skip > 0:
            current = doc_date + relativedelta(months=interval * skip)
    while current < target_end:
        if target_start <= current < target_end:
            return current.strftime("%d/%m/%Y")
        current += relativedelta(months=interval)
        if current.year > year + 2: break
    return None

AUTO_STATUSES = {"AUTODEBIT", "BRANCHPAID", "DAILYCOLLECTION"}

def normalize_status(s):
    if not s: return ""
    raw = str(s).strip()
    if not raw: return ""
    c = re.sub(r"[^a-z0-9]", "", raw.lower())
    if not c: return ""
    # AUTODEBIT variations: auto, auto debit, ad, a/d, ecs, nach, si, autopay, etc.
    auto_matches = (
        "autodebit", "auto", "ad", "ecs", "nach", "si",
        "autopay", "autop", "standinginstruction", "emandate",
        "mandate", "neft", "autopremium", "ap",
    )
    if c in auto_matches or c.startswith("auto") or c.startswith("nach") or c.startswith("ecs"):
        return "AUTODEBIT"
    # BRANCHPAID variations: branch, branch paid, br, bp, etc.
    branch_matches = (
        "branchpaid", "branch", "br", "bp", "branchcollection",
        "branchcoll", "brpaid",
    )
    if c in branch_matches or c.startswith("branch"):
        return "BRANCHPAID"
    # DAILYCOLLECTION variations: daily, daily collection, dc, d/c, etc.
    daily_matches = (
        "dailycollection", "daily", "dc", "dailycoll",
        "dailypayment", "dailyprem",
    )
    if c in daily_matches or c.startswith("daily"):
        return "DAILYCOLLECTION"
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  PERSISTENCE LAYER  — read/write JSON blobs to kv_store
# ══════════════════════════════════════════════════════════════════════════════

def persist_sheet(key):
    """Write one sheet from cache to kv_store (background, non-blocking)."""
    try:
        with _cache_lock:
            sheet = CACHE["sheets"].get(key)
        if sheet is None: return
        data = json.dumps(sheet, ensure_ascii=False)
        db_exec("INSERT OR REPLACE INTO kv_store (key, value) VALUES (?, ?)",
                (f"sheet_{key}", data))
    except Exception as e:
        log.error(f"[persist_sheet] {key}: {e}")

def persist_col_names():
    """Write column names to kv_store."""
    try:
        data = json.dumps(CACHE["col_names"], ensure_ascii=False)
        db_exec("INSERT OR REPLACE INTO kv_store (key, value) VALUES (?, ?)",
                ("col_names", data))
    except Exception as e:
        log.error(f"[persist_col_names] {e}")

def persist_meta():
    """Write id counter to kv_store."""
    try:
        db_exec("INSERT OR REPLACE INTO kv_store (key, value) VALUES (?, ?)",
                ("_id_counter", str(CACHE["_id_counter"])))
    except Exception as e:
        log.error(f"[persist_meta] {e}")

def delete_sheet_from_store(key):
    """Remove sheet from kv_store."""
    try:
        db_exec("DELETE FROM kv_store WHERE key=?", (f"sheet_{key}",))
    except Exception as e:
        log.error(f"[delete_sheet] {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  STARTUP — init tables + hydrate cache from Turso
# ══════════════════════════════════════════════════════════════════════════════

def init_db():
    """Create tables and hydrate in-memory cache."""
    # Create tables
    db_exec("CREATE TABLE IF NOT EXISTS master_policies ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "policyno TEXT UNIQUE NOT NULL,"
            "name TEXT DEFAULT '', doc TEXT DEFAULT '', plan TEXT DEFAULT '',"
            "mode TEXT DEFAULT '', premium TEXT DEFAULT '', sumass TEXT DEFAULT '',"
            "mobileno TEXT DEFAULT '', status TEXT DEFAULT '',"
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
            "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")

    db_exec("CREATE TABLE IF NOT EXISTS kv_store ("
            "key TEXT PRIMARY KEY, value TEXT)")

    # Hydrate cache — load everything in ONE pipeline call
    today = date.today()
    months_to_load = []
    for delta in range(3):  # current + past 2 months
        d = date(today.year, today.month, 1) - relativedelta(months=delta)
        months_to_load.append((d.year, d.month))

    queries = [
        ("SELECT COUNT(*) AS c FROM master_policies", None),
        ("SELECT key, value FROM kv_store", None),
    ]
    results = db_pipeline(queries)

    # Master count
    CACHE["master_count"] = results[0][0]["c"] if results[0] else 0

    # Load all KV data
    kv_data = {r["key"]: r["value"] for r in results[1]} if len(results) > 1 else {}

    # Restore column names
    if "col_names" in kv_data:
        try:
            CACHE["col_names"] = json.loads(kv_data["col_names"])
        except:
            pass

    # Restore ID counter
    if "_id_counter" in kv_data:
        try:
            CACHE["_id_counter"] = int(kv_data["_id_counter"])
        except:
            pass

    # Restore sheets from kv_store
    for k, v in kv_data.items():
        if k.startswith("sheet_"):
            try:
                sheet = json.loads(v)
                cache_key = k[6:]  # remove "sheet_" prefix
                CACHE["sheets"][cache_key] = sheet
            except:
                pass

    # Fill default col names
    for i in range(1, COL_EXTRA + 1):
        if f"col{i}" not in CACHE["col_names"]:
            CACHE["col_names"][f"col{i}"] = f"Note {i}"

    CACHE["_loaded"] = True
    print(f"[Startup] Cache hydrated: {CACHE['master_count']} master, "
          f"{len(CACHE['sheets'])} sheets, id_counter={CACHE['_id_counter']}")

try:
    init_db()
    print(f"[Startup] OK — USE_TURSO={USE_TURSO}")
except Exception as e:
    print(f"[Startup] FAILED: {e}")
    CACHE["_loaded"] = True  # allow app to start even if DB fails


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def generate_sheet(year, month):
    """Compute FUP sheet from master data → save to cache + persist."""
    mc = db_exec("SELECT COUNT(*) AS c FROM master_policies")
    if not mc or mc[0]["c"] == 0:
        raise HTTPException(400, "No master data. Upload policies first.")

    all_policies = db_exec("SELECT * FROM master_policies ORDER BY id")

    due = []
    for p in all_policies:
        d, m = p.get("doc", ""), p.get("mode", "")
        if d and m and is_due_in_month(d, m, year, month):
            dd = get_due_date(d, m, year, month)
            ms = normalize_status(p.get("status", ""))
            entry_id = next_id()
            due.append({
                "id": entry_id,
                "sn": len(due) + 1,
                "policyno": p.get("policyno", ""),
                "name": p.get("name", ""),
                "doc": p.get("doc", ""),
                "plan": p.get("plan", ""),
                "mode": p.get("mode", ""),
                "premium": p.get("premium", ""),
                "sumass": p.get("sumass", ""),
                "mobileno": p.get("mobileno", ""),
                "due_date": dd or "",
                "status": ms or "DUE",
                **{f"col{i}": "" for i in range(1, COL_EXTRA + 1)},
            })

    key = sheet_key(year, month)
    sheet_obj = {
        "month": month,
        "year": year,
        "generated_at": datetime.now().isoformat(),
        "total_count": len(due),
        "entries": due,
    }

    with _cache_lock:
        CACHE["sheets"][key] = sheet_obj

    # Persist in background thread (non-blocking)
    threading.Thread(target=persist_sheet, args=(key,), daemon=True).start()
    threading.Thread(target=persist_meta, daemon=True).start()

    return {"sheet_id": key, "month": month, "year": year,
            "total_count": len(due), "generated_at": sheet_obj["generated_at"]}


# ── Background Scheduler ──────────────────────────────────────────────────────

def auto_generate_current_month():
    today = date.today()
    key = sheet_key(today.year, today.month)
    try:
        if CACHE["master_count"] == 0:
            return
        if key not in CACHE["sheets"]:
            generate_sheet(today.year, today.month)
            print(f"[AutoGen] Generated {today.strftime('%B %Y')}")
    except Exception as e:
        print(f"[AutoGen] Error: {e}")

def self_ping():
    try:
        port = int(os.environ.get("PORT", 8900))
        http_requests.get(f"http://127.0.0.1:{port}/health", timeout=5)
    except:
        pass

scheduler = BackgroundScheduler()
scheduler.add_job(auto_generate_current_month, 'cron', day=1, hour=0, minute=5)
scheduler.add_job(auto_generate_current_month, 'date',
                  run_date=datetime.now() + timedelta(seconds=10))
scheduler.add_job(self_ping, 'interval', minutes=10)
scheduler.start()


# ══════════════════════════════════════════════════════════════════════════════
#  API ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/init/{year}/{month}")
def api_init(year: int, month: int):
    """⚡ Serves ENTIRELY from in-memory cache. Zero DB calls."""
    today = date.today()
    prev = date(today.year, today.month, 1) - timedelta(days=1)
    pm, py_ = prev.month, prev.year

    key = sheet_key(year, month)
    sheet = CACHE["sheets"].get(key)
    prev_key = sheet_key(py_, pm)
    prev_sheet = CACHE["sheets"].get(prev_key)

    unpaid_entries = []
    if prev_sheet:
        unpaid_entries = [e for e in prev_sheet.get("entries", []) if e.get("status") == "DUE"]

    return {
        "master_count": CACHE["master_count"],
        "col_names": CACHE["col_names"],
        "sheet": {
            "exists": sheet is not None,
            **(sheet if sheet else {"month": month, "year": year}),
            "entries": sheet["entries"] if sheet else [],
        },
        "unpaid": {
            "has_unpaid": len(unpaid_entries) > 0,
            "count": len(unpaid_entries),
            "month": pm, "year": py_,
            "entries": unpaid_entries,
        },
    }


# ── Master Data Endpoints ─────────────────────────────────────────────────────

@app.post("/api/master/upload")
async def api_upload_master(file: UploadFile = File(...)):
    content = await file.read()
    fn = (file.filename or "").lower()
    if fn.endswith('.xlsx') or fn.endswith('.xlsm'):
        rows = parse_excel_file(content)
    elif fn.endswith('.xls'):
        rows = parse_xls_file(content)
    elif fn.endswith('.csv'):
        rows = parse_csv_file(content)
    else:
        raise HTTPException(400, "Use .csv, .xlsx, .xlsm, or .xls files only.")

    if len(rows) < 2:
        raise HTTPException(400, "File needs a header row + at least one data row.")

    col_map = map_columns(rows[0])
    if 'policyno' not in col_map.values():
        raise HTTPException(400, f"Cannot detect Policy No column. Headers: {rows[0]}")

    # Get existing policy numbers
    existing = {r["policyno"] for r in db_exec("SELECT policyno FROM master_policies")}
    new_c = upd_c = 0
    now = datetime.now().isoformat()

    stmts = []
    for row in rows[1:]:
        data = {}
        for idx, field in col_map.items():
            if idx < len(row):
                data[field] = str(row[idx]).strip() if row[idx] else ""
        pno = data.get('policyno', '').strip().upper()
        if not pno: continue
        st = normalize_status(data.get('status', ''))
        if pno in existing:
            stmts.append((
                "UPDATE master_policies SET name=?,doc=?,plan=?,mode=?,premium=?,sumass=?,mobileno=?,status=?,updated_at=? WHERE policyno=?",
                (data.get('name',''), data.get('doc',''), data.get('plan',''),
                 data.get('mode',''), data.get('premium',''), data.get('sumass',''),
                 data.get('mobileno',''), st, now, pno)))
            upd_c += 1
        else:
            stmts.append((
                "INSERT INTO master_policies (policyno,name,doc,plan,mode,premium,sumass,mobileno,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (pno, data.get('name',''), data.get('doc',''), data.get('plan',''),
                 data.get('mode',''), data.get('premium',''), data.get('sumass',''),
                 data.get('mobileno',''), st, now, now)))
            existing.add(pno)
            new_c += 1

    if stmts:
        db_batch(stmts)

    total = db_exec("SELECT COUNT(*) AS c FROM master_policies")
    CACHE["master_count"] = total[0]["c"] if total else 0

    return {"new": new_c, "updated": upd_c,
            "total_master": CACHE["master_count"],
            "columns_detected": {rows[0][k]: v for k, v in col_map.items() if k < len(rows[0])}}


@app.post("/api/master/single")
def api_add_single(policyno: str = Query(...), name: str = Query(""),
    doc: str = Query(""), plan: str = Query(""), mode: str = Query(""),
    premium: str = Query(""), sumass: str = Query(""), mobileno: str = Query(""),
    status: str = Query("")):
    pno = policyno.strip().upper()
    if not pno: raise HTTPException(400, "Policy number required")
    st = normalize_status(status)
    now = datetime.now().isoformat()
    ex = db_exec("SELECT id FROM master_policies WHERE policyno=?", (pno,))
    if ex:
        db_exec(
            "UPDATE master_policies SET name=?,doc=?,plan=?,mode=?,premium=?,sumass=?,mobileno=?,status=?,updated_at=? WHERE policyno=?",
            (name.strip(), doc.strip(), plan.strip(), mode.strip(), premium.strip(),
             sumass.strip(), mobileno.strip(), st, now, pno))
        return {"action": "updated", "policyno": pno}
    db_exec(
        "INSERT INTO master_policies (policyno,name,doc,plan,mode,premium,sumass,mobileno,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (pno, name.strip(), doc.strip(), plan.strip(), mode.strip(), premium.strip(),
         sumass.strip(), mobileno.strip(), st, now, now))
    mc = db_exec("SELECT COUNT(*) AS c FROM master_policies")
    CACHE["master_count"] = mc[0]["c"] if mc else 0
    return {"action": "added", "policyno": pno}


@app.get("/api/master/count")
def api_master_count():
    return {"count": CACHE["master_count"]}


@app.get("/api/master/status-check")
def api_status_check():
    """Diagnostic: show raw status values in master data and what they normalize to."""
    rows = db_exec("SELECT status, COUNT(*) as cnt FROM master_policies GROUP BY status ORDER BY cnt DESC")
    result = []
    for r in rows:
        raw = r.get("status", "")
        result.append({
            "raw_status": raw,
            "normalized_to": normalize_status(raw) or "(empty → DUE)",
            "count": r["cnt"],
        })
    return {"status_breakdown": result, "total": sum(r["cnt"] for r in result)}



@app.delete("/api/master/reset")
def api_reset_master():
    db_exec("DELETE FROM master_policies")
    CACHE["master_count"] = 0
    return {"message": "All master data deleted."}


# ── Sheet Endpoints ────────────────────────────────────────────────────────────

@app.post("/api/sheets/generate/{year}/{month}")
def api_generate_sheet(year: int, month: int):
    return generate_sheet(year, month)


@app.get("/api/sheets/{year}/{month}")
def api_get_sheet(year: int, month: int):
    key = sheet_key(year, month)
    sheet = CACHE["sheets"].get(key)
    if not sheet:
        return {"exists": False, "month": month, "year": year, "entries": []}
    return {"exists": True, **sheet}


@app.delete("/api/sheets/{year}/{month}")
def api_delete_sheet(year: int, month: int):
    key = sheet_key(year, month)
    if key not in CACHE["sheets"]:
        raise HTTPException(404, "Sheet not found")
    with _cache_lock:
        del CACHE["sheets"][key]
    threading.Thread(target=delete_sheet_from_store, args=(key,), daemon=True).start()
    return {"message": f"Sheet {month}/{year} deleted."}


@app.post("/api/sheets/{year}/{month}/entry")
def api_add_entry(year: int, month: int,
    policyno: str = Query(""), name: str = Query(""), doc: str = Query(""),
    plan: str = Query(""), mode: str = Query(""), premium: str = Query(""),
    sumass: str = Query(""), mobileno: str = Query(""), due_date: str = Query("")):

    key = sheet_key(year, month)
    with _cache_lock:
        sheet = CACHE["sheets"].get(key)
        if not sheet:
            sheet = {"month": month, "year": year, "generated_at": datetime.now().isoformat(),
                     "total_count": 0, "entries": []}
            CACHE["sheets"][key] = sheet

        entry_id = next_id()
        nsn = len(sheet["entries"]) + 1
        entry = {
            "id": entry_id, "sn": nsn,
            "policyno": policyno.strip(), "name": name.strip(),
            "doc": doc.strip(), "plan": plan.strip(), "mode": mode.strip(),
            "premium": premium.strip(), "sumass": sumass.strip(),
            "mobileno": mobileno.strip(), "due_date": due_date.strip(),
            "status": "DUE",
            **{f"col{i}": "" for i in range(1, COL_EXTRA + 1)},
        }
        sheet["entries"].append(entry)
        sheet["total_count"] = len(sheet["entries"])

    threading.Thread(target=persist_sheet, args=(key,), daemon=True).start()
    threading.Thread(target=persist_meta, daemon=True).start()
    return {"entry": entry}


# ── Entry Edit Endpoints (cache-first, async persist) ─────────────────────────

@app.patch("/api/entries/{entry_id}/toggle")
def api_toggle_status(entry_id: int):
    # Find entry in cache
    for key, sheet in CACHE["sheets"].items():
        for entry in sheet["entries"]:
            if entry["id"] == entry_id:
                cur = entry["status"]
                if cur in AUTO_STATUSES:
                    raise HTTPException(400, "Cannot toggle auto-payment statuses.")
                new = "PAID" if cur == "DUE" else "DUE"
                entry["status"] = new
                threading.Thread(target=persist_sheet, args=(key,), daemon=True).start()
                return {"id": entry_id, "status": new}
    raise HTTPException(404, "Not found")


@app.patch("/api/entries/{entry_id}/cell")
def api_update_cell(entry_id: int, col: str = Query(...), value: str = Query("")):
    allowed = {"sn","policyno","name","doc","plan","mode","premium","sumass","mobileno","due_date"}
    allowed.update({f"col{i}" for i in range(1, COL_EXTRA + 1)})
    if col not in allowed:
        raise HTTPException(400, f"Invalid column: {col}")

    for key, sheet in CACHE["sheets"].items():
        for entry in sheet["entries"]:
            if entry["id"] == entry_id:
                entry[col] = value
                threading.Thread(target=persist_sheet, args=(key,), daemon=True).start()
                return {"id": entry_id, "col": col, "value": value}
    raise HTTPException(404, "Not found")


# ── Column Names ──────────────────────────────────────────────────────────────

@app.get("/api/col-names")
def api_get_col_names():
    return CACHE["col_names"]

@app.patch("/api/col-names")
def api_set_col_name(col: str = Query(...), name: str = Query(...)):
    if col not in {f"col{i}" for i in range(1, COL_EXTRA + 1)}:
        raise HTTPException(400, "Invalid column")
    CACHE["col_names"][col] = name.strip()
    threading.Thread(target=persist_col_names, daemon=True).start()
    return {"col": col, "name": name.strip()}


# ── Export ────────────────────────────────────────────────────────────────────

@app.get("/api/sheets/{year}/{month}/csv")
def api_export_csv(year: int, month: int):
    key = sheet_key(year, month)
    sheet = CACHE["sheets"].get(key)
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    out = io.StringIO()
    w = csv.writer(out)
    h = ["SN","Policy No","Name","DOC","Plan","Mode","Premium","Sum Assured","Mobile","Due Date","Status"]
    h += [CACHE["col_names"].get(f"col{i}", f"Note {i}") for i in range(1, COL_EXTRA + 1)]
    w.writerow(h)
    for e in sheet["entries"]:
        r = [e.get("sn",""), e.get("policyno",""), e.get("name",""), e.get("doc",""),
             e.get("plan",""), e.get("mode",""), e.get("premium",""), e.get("sumass",""),
             e.get("mobileno",""), e.get("due_date",""), e.get("status","")]
        r += [e.get(f"col{i}", "") for i in range(1, COL_EXTRA + 1)]
        w.writerow(r)
    out.seek(0)
    MN = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="DueList_{MN[month]}_{year}.csv"'})


# ── Unpaid Alert ──────────────────────────────────────────────────────────────

@app.get("/api/unpaid-alert")
def api_unpaid_alert():
    today = date.today()
    prev = date(today.year, today.month, 1) - timedelta(days=1)
    pm, py_ = prev.month, prev.year
    key = sheet_key(py_, pm)
    sheet = CACHE["sheets"].get(key)
    if not sheet:
        return {"has_unpaid": False, "count": 0, "month": pm, "year": py_}
    ue = [e for e in sheet.get("entries", []) if e.get("status") == "DUE"]
    return {"has_unpaid": len(ue) > 0, "count": len(ue), "month": pm, "year": py_,
            "entries": ue}


# ── Misc ──────────────────────────────────────────────────────────────────────

@app.get("/api/time")
def api_time():
    return {"time": datetime.now().isoformat()}

@app.get("/")
def root():
    return HTMLResponse(open(os.path.join(BASE_DIR, "index.html"), encoding="utf-8").read())

@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    return {"status": "ok", "version": "2.1.0"}

@app.get("/api/debug")
def api_debug():
    return {
        "USE_TURSO": USE_TURSO,
        "version": "2.1.0",
        "master": CACHE["master_count"],
        "sheets_cached": list(CACHE["sheets"].keys()),
        "sheets_count": len(CACHE["sheets"]),
        "total_entries": sum(len(s.get("entries", [])) for s in CACHE["sheets"].values()),
        "id_counter": CACHE["_id_counter"],
        "cache_loaded": CACHE["_loaded"],
    }
